# Django
from django.contrib import auth
from django.core.files import File
from django.db.models.aggregates import Min, Max, Count
from django.db.models import FloatField, F, Q, Case, When, Value, IntegerField, CharField, ExpressionWrapper, Func
from django.db.models.functions import Cast, ExtractMonth, ExtractDay, Now, Round, Substr, ExtractYear
from django.forms import DurationField
from django.utils import timezone

# Utilities
from dashboards.models import Aplicacion, Bitacora, Compania, Inspeccion, Llanta, Perfil, Ubicacion, Vehiculo, Producto, FTP
from datetime import date, datetime, timedelta
from heapq import nlargest
from itertools import count
import math
import matplotlib.pyplot as plt 
import numpy as np
from random import sample, randint, uniform, random
from sklearn.metrics import r2_score
from openpyxl import load_workbook
import pandas as pd
import os
import statistics

class DiffDays(Func):
    function = 'DATE_PART'
    template = "%(function)s('day', %(expressions)s)"

class CastDate(Func):
    function = 'date_trunc'
    template = "%(function)s('day', %(expressions)s)"



def crear_nombre_de_eje():
    """llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania="Compania Prueba"))
    for llanta in llantas:
        num_aleatorio = sorted(sample(range(2, 5), 2))
        i = Inspeccion.objects.create(
            llanta=llanta,
            fecha_hora=datetime.today(),
            tiempo_de_inspeccion=2,
            min_profundidad=num_aleatorio[0],
            max_profundidad=num_aleatorio[1],
        )
        llanta.ultima_inspeccion = i
        llanta.save()"""

def crear_1(numero_economico, compania, ubicacion, aplicacion, clase, configuracion, tiempo_de_inflado, presion_de_entrada, presion_de_salida, presion_establecida):
    """Vehiculo.objects.create(numero_economico=numero_economico,
                                modelo="Kenworth",
                                marca="International",
                                compania=Compania.objects.get(compania=compania),
                                ubicacion=Ubicacion.objects.get(nombre=ubicacion),
                                aplicacion=Aplicacion.objects.get(nombre=aplicacion),
                                clase=clase,
                                configuracion=configuracion,
                                fecha_de_inflado=date.today(),
                                tiempo_de_inflado=tiempo_de_inflado,
                                presion_de_entrada=presion_de_entrada,
                                presion_de_salida=presion_de_salida,
                                presion_establecida=presion_establecida
                            )
    Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=numero_economico),
        compania=Compania.objects.get(compania=Compania.objects.get(compania=compania)),
        fecha_de_inflado=date.today(),
        tiempo_de_inflado=tiempo_de_inflado,
        presion_de_entrada=presion_de_entrada,
        presion_de_salida=presion_de_salida,
        presion_establecida=presion_establecida
        )"""
        
def crear_2(vehiculo):
    """num_llantas = 0
    for i in vehiculo.configuracion:
        try:
            i = int(i)
            num_llantas += i
        except:
            pass
    iteracion_ejes = {}
    ejes = vehiculo.configuracion.split(".")
    iteracion_ejes[1] = [ejes[0], int(ejes[0][1])]
    iteracion_ejes[2] = [ejes[1], int(ejes[1][1])]
    try:
        iteracion_ejes[ejes[2]] = [ejes[2], int(ejes[2][1])]
    except:
        pass
    familia_ejes = 0
    iteracion_familia_ejes = 0
    posiciones = ["LI", "RI", "LO", "RO"]
    for i in range(1, num_llantas + 1):            
        Llanta.objects.create(numero_economico=f"{vehiculo}-{i}",
            usuario=Perfil.objects.get(user__username="pruebacal"),
            vehiculo=vehiculo,
            producto="NUEVA Michelin XDA5+ LRH 11R22.5",
            marca="Michelin",
            tipo_de_eje=iteracion_ejes[familia_ejes + 1][0],
            eje=familia_ejes + 1,
            posicion=f"{familia_ejes + 1}{posiciones[iteracion_familia_ejes]}",
            fecha_de_inflado=date.today()
        )
        iteracion_familia_ejes += 1
        if iteracion_familia_ejes == iteracion_ejes[familia_ejes + 1][1]:
            familia_ejes += 1
            iteracion_familia_ejes = 0"""

def crear_3(vehiculo):
    """llantas = Llanta.objects.filter(vehiculo=vehiculo)
    for llanta in llantas:
        num_aleatorio = sorted(sample(range(2, 10), 2))

        i = Inspeccion.objects.create(
            llanta=llanta,
            fecha_hora=datetime(2021, 11, 20, 12, 0, 1),
            tiempo_de_inspeccion=2,
            min_profundidad=num_aleatorio[0],
            max_profundidad=num_aleatorio[1],
        )
        llanta.ultima_inspeccion = i
        llanta.save()"""

def crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes, posicion, user):
    """presion_encontrada = randint(80, 100)
    tiempo_de_inspeccion = round(uniform(2, 3), 1)
    profundidad = [randint(2, 8), randint(2, 8)]
    primera_profundidad = [randint(2, 8), randint(2, 8)]
    profundidad.sort()
    primera_profundidad.sort()
    inicio = datetime(2021, 8, 1)
    final =  datetime(2021, 11, 3)
    fecha_hora_de_inspeccion = inicio + (final - inicio) * random()

    vehiculo.fecha_hora_de_inspeccion = fecha_hora_de_inspeccion
    vehiculo.save()

    Llanta.objects.create(numero_economico=f"AETO-LLANTA-{num_aleatorio[iteracion_llantas_totales]}",
                            usuario=Perfil.objects.get(user=user),
                            vehiculo=vehiculo,
                            producto="Bridgstone A1",
                            marca="Freightliner",
                            tipo_de_eje=eje,
                            eje=iteracion_ejes,
                            posicion=posicion,
                            presion_encontrada=presion_encontrada,
                            presion_actual=vehiculo.presion_de_entrada,
                            presion_establecida=vehiculo.presion_establecida,
                            fecha_de_inflado=vehiculo.fecha_de_inflado,
                            fecha_hora_de_inspeccion=fecha_hora_de_inspeccion,
                            tiempo_de_inspeccion=tiempo_de_inspeccion,
                            primera_min_profundidad=primera_profundidad[0],
                            primera_max_profundidad=primera_profundidad[1],
                            min_profundidad=profundidad[0],
                            max_profundidad=profundidad[1],
                            )"""

def crear_llanta_sin_vehiculo(user):

    for i in range(1, 51):
        tiempo_de_inspeccion = round(uniform(2, 3), 1)
        profundidad = [randint(2, 8), randint(2, 8)]
        primera_profundidad = [randint(2, 8), randint(2, 8)]
        profundidad.sort()
        primera_profundidad.sort()
        inicio = datetime(2021, 8, 1)
        final =  datetime(2021, 11, 3)
        fecha_hora_de_inspeccion = inicio + (final - inicio) * random()
        producto = randint(1, 3)
        vida = randint(1, 3)
        inventario = randint(1, 6)

        if vida == 1:
            vida = "Nueva"
        elif vida == 2:
            vida = "1R"
        elif vida == 3:
            vida = "2R"

        if vida == "Nueva":
            inventario = "Nueva"
        else:
            if inventario == 1:
                inventario = "Antes de Renovar"
            if inventario == 2:
                inventario = "Antes de Desechar"
            if inventario == 3:
                inventario = "Renovada"
            if inventario == 4:
                inventario = "Con renovador"
            if inventario == 5:
                inventario = "Desecho Final"
            if inventario == 6:
                inventario = "Servicio"


        Llanta.objects.create(numero_economico=f"PruebaCal-{i}",
                                usuario=Perfil.objects.get(user=user),
                                producto=Producto.objects.get(id=producto),
                                vida=vida,
                                inventario=inventario
                                )

def crear_producto():
    """llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania="pruebacal"))
    for llanta in llantas:

        if llanta.numero_economico[0:2] == "P1":
            llanta.nombre_de_eje = "Arrastre"
            llanta.save()
        if llanta.numero_economico[0:2] == "P2":
            llanta.nombre_de_eje = "Arrastre"
            llanta.save()
        if llanta.numero_economico[0:2] == "P3":
            llanta.nombre_de_eje = "Arrastre"
            llanta.save()
        if llanta.numero_economico[0:2] == "P4":
            llanta.nombre_de_eje = "Arrastre"
            llanta.save()"""

def crear_mm():
    """inspecciones = Inspeccion.objects.filter(llanta__vehiculo__compania=Compania.objects.get(compania="pruebacal"))
    for inspeccion in inspecciones:
        num_aleatorio = sorted(sample(range(4, 10), 2))
        
        inspeccion.min_profundidad = num_aleatorio[0]
        inspeccion.max_profundidad = num_aleatorio[1]
        inspeccion.save()"""

def crear_fecha():
    """inspecciones = Inspeccion.objects.filter(llanta__vehiculo__compania=Compania.objects.get(compania="pruebacal"))
    llantas_repetidas = []
    for inspeccion in inspecciones:
        llanta = str(inspeccion)[0:4]
        if not(llanta in llantas_repetidas):
            llantas_repetidas.append(llanta)
        else:
            inspeccion.km = 15000
            inspeccion.save()"""
        

def excel(user):
    
    """num_aleatorio = sample(range(1, 100000), 36000)
    vehiculos = Vehiculo.objects.filter(compania=Compania.objects.get(compania="Compania Prueba"))
    iteracion_vehiculo = 0
    iteracion_llantas_totales = 0
    iteracion_configuraciones = ["S2.D2", "S2.D2.D2", "S2.D4", "S2.D2.D2.T4.T4"]
    for vehiculo in vehiculos:
        iteracion_ejes = {}
        if iteracion_vehiculo < 400:
            vehiculo.configuracion = iteracion_configuraciones[0]
            vehiculo.save()
            ejes = iteracion_configuraciones[0].split(".")
            iteracion_ejes[ejes[0]] = [int(ejes[0][1]), 1]
            iteracion_ejes[ejes[1]] = [int(ejes[1][1]), 2]
        elif iteracion_vehiculo < 1800:
            vehiculo.configuracion = iteracion_configuraciones[1]
            vehiculo.save()
            ejes = iteracion_configuraciones[1].split(".")
            iteracion_ejes[ejes[0]] = [int(ejes[0][1]), 1]
            iteracion_ejes[ejes[1]] = [int(ejes[1][1]), 2]
            iteracion_ejes[ejes[2]] = [int(ejes[2][1]), 3]
        elif iteracion_vehiculo < 2600:
            vehiculo.configuracion = iteracion_configuraciones[2]
            vehiculo.save()
            ejes = iteracion_configuraciones[2].split(".")
            iteracion_ejes[ejes[0]] = [int(ejes[0][1]), 1]
            iteracion_ejes[ejes[1]] = [int(ejes[1][1]), 2]
        else:
            vehiculo.configuracion = iteracion_configuraciones[3]
            vehiculo.save()
            ejes = iteracion_configuraciones[3].split(".")
            iteracion_ejes[ejes[0]] = [int(ejes[0][1]), 1]
            iteracion_ejes[ejes[1]] = [int(ejes[1][1]), 2]
            iteracion_ejes[ejes[2]] = [int(ejes[2][1]), 3]
            iteracion_ejes[ejes[3]] = [int(ejes[3][1]), 4]

        for eje in iteracion_ejes:

            if iteracion_ejes[eje][0] == 2:
                posicion = f"{iteracion_ejes[eje][1]}RI"
                crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes[eje][1], posicion, user)
                iteracion_llantas_totales += 1
                posicion = f"{iteracion_ejes[eje][1]}LI"
                crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes[eje][1], posicion, user)
                iteracion_llantas_totales += 1
            elif iteracion_ejes[eje][0] == 4:
                posicion = f"{iteracion_ejes[eje][1]}RI"
                crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes[eje][1], posicion, user)
                iteracion_llantas_totales += 1
                posicion = f"{iteracion_ejes[eje][1]}RO"
                crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes[eje][1], posicion, user)
                iteracion_llantas_totales += 1
                posicion = f"{iteracion_ejes[eje][1]}LI"
                crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes[eje][1], posicion, user)
                iteracion_llantas_totales += 1
                posicion = f"{iteracion_ejes[eje][1]}LO"
                crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes[eje][1], posicion, user)
                iteracion_llantas_totales += 1
        iteracion_vehiculo += 1"""

    """num_aleatorio = sample(range(1, 100000), 3000)
    iteracion = 0
    for num in num_aleatorio:
        iteracion += 1
        if iteracion <= 400:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Kenworth",
                                        marca="International",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="CAMIONETA",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Camionera Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.4,
                                        presion_de_entrada=100,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.4,
                presion_de_entrada=100,
                presion_de_salida=100,
                )
        elif iteracion <= 500:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Tesla",
                                        marca="International",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="AUTOTANQUE ALIMENTICIO",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Camionera Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.3,
                                        presion_de_entrada=100,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.4,
                presion_de_entrada=100,
                presion_de_salida=100,
            )
        elif iteracion <= 640:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Tesla",
                                        marca="International",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="AUTOTANQUE ALIMENTICIO",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Camionera Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.0,
                                        presion_de_entrada=75,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.0,
                presion_de_entrada=75,
                presion_de_salida=100,
            )
        elif iteracion <= 700:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="AUTOTANQUE ALIMENTICIO",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Camionera Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.0,
                                        presion_de_entrada=100,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.0,
                presion_de_entrada=100,
                presion_de_salida=100,
            )
        elif iteracion <= 720:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="AUTOTANQUE ALIMENTICIO",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Camionera Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.8,
                                        presion_de_entrada=80,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.8,
                presion_de_entrada=80,
                presion_de_salida=100,
            )
        elif iteracion <= 720:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="AUTOTANQUE ALIMENTICIO",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Transportes Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.7,
                                        presion_de_entrada=82,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.7,
                presion_de_entrada=82,
                presion_de_salida=100,
            )
        elif iteracion <= 790:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="AUTOTANQUE ALIMENTICIO",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Transportes Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.5,
                                        presion_de_entrada=100,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.5,
                presion_de_entrada=100,
                presion_de_salida=100,
            )
        elif iteracion <= 1200:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="REMOLQUE",
                                        ubicacion=Ubicacion.objects.get(nombre="Monterrey Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Transportes Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.5,
                                        presion_de_entrada=100,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.5,
                presion_de_entrada=100,
                presion_de_salida=100,
            )
        elif iteracion <= 2200:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="TOLVA",
                                        ubicacion=Ubicacion.objects.get(nombre="Monterrey Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Transportes Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.2,
                                        presion_de_entrada=100,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.2,
                presion_de_entrada=100,
                presion_de_salida=100,
            )
        elif iteracion <= 3000:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="TOLVA",
                                        ubicacion=Ubicacion.objects.get(nombre="Monterrey Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Camionera Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.6,
                                        presion_de_entrada=72,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.6,
                presion_de_entrada=72,
                presion_de_salida=100,
            )"""
    
    """FILE_PATH = "D:/Descargas/Tablas AETO Tire.xlsx"
    SHEET = "Inspeccion"
    workbook = load_workbook(FILE_PATH, read_only=True)
    sheet = workbook[SHEET]

    for row in sheet.iter_rows(min_row=2):
        numero_economico_de_llanta = row[1].value
        min_profundidad = int(row[4].value)

        Inspeccion.objects.create(llanta=Llanta.objects.get(numero_economico=numero_economico_de_llanta),
                            fecha_hora=date.today(),
                            tiempo_de_inspeccion=1,
                            min_profundidad=min_profundidad,
                            max_profundidad=min_profundidad,
                            )"""

    """while vehiculos_chidos.count():
        ids = Vehiculo.objects.values_list('pk', flat=True)[:100]
        Vehiculo.objects.filter(pk__in = ids, compania=Compania.objects.get(compania="Idealease")).delete()
    """



def clases_mas_frecuentes(vehiculo_fecha, compania):
    try:
        vehiculos = Vehiculo.objects.filter(compania=compania)
        clases_compania = []
        for vehiculo in vehiculos:
            if not(vehiculo.clase in clases_compania):
                clases_compania.append(vehiculo.clase.capitalize())
        clases = {}
        for v in vehiculo_fecha:
            clase = v.clase.capitalize()
            if clase in clases:
                clases[clase] += 1
            else:
                clases[clase] = 1
        if 2 < len(clases) <= 8:
            clases_mayores = nlargest(len(clases), clases, key=clases.get)
        elif 0 <= len(clases) <= 2:
            for clase_compania in clases_compania:
                if not(clase_compania in clases):
                    clases[clase_compania] = 0
            clases_mayores = nlargest(8, clases, key=clases.get)
        else:
            clases_mayores = nlargest(8, clases, key=clases.get)
        for c in clases.copy():
            if c in clases_mayores:
                vehiculo_clase = Vehiculo.objects.filter(clase__icontains=c.upper(), compania=compania)
                clases[c] = round((clases[c] / vehiculo_clase.count()) * 100, 2)
            else:
                clases.pop(c)
        return clases
    except:
        return None

def aplicaciones_mas_frecuentes(vehiculo_fecha, vehiculos, compania):
    try:
        aplicaciones_compania = []
        for vehiculo in vehiculos:
            if not(vehiculo.aplicacion in aplicaciones_compania):
                aplicaciones_compania.append(vehiculo.aplicacion)
        aplicaciones = {}
        for v in vehiculo_fecha:
            aplicacion = v.aplicacion
            if aplicacion in aplicaciones:
                aplicaciones[aplicacion] += 1
            else:
                aplicaciones[aplicacion] = 1
        if 2 < len(aplicaciones) <= 8:
            aplicaciones_mayores = nlargest(len(aplicaciones), aplicaciones, key=aplicaciones.get)
        elif 0 <= len(aplicaciones) <= 2:
            for aplicacion_compania in aplicaciones_compania:
                if not(aplicacion_compania in aplicaciones):
                    aplicaciones[aplicacion_compania] = 0
            aplicaciones_mayores = nlargest(8, aplicaciones, key=aplicaciones.get)
        else:
            aplicaciones_mayores = nlargest(8, aplicaciones, key=aplicaciones.get)
        for c in aplicaciones.copy():
            if c in aplicaciones_mayores:
                vehiculo_aplicacion = vehiculos.filter(aplicacion=c, compania=compania)
                aplicaciones[c] = round((aplicaciones[c] / vehiculo_aplicacion.count()) * 100, 2)
            else:
                aplicaciones.pop(c)
        return aplicaciones
    except:
        return None

def cantidad_llantas(configuracion):
    llantas = 0
    for caracter in configuracion:
        if caracter.isdigit():
            llantas += int(caracter)
    return llantas


def contar_dias(fecha):
    fecha_date = datetime.strptime(fecha, "%Y-%m-%d").date()
    hoy = date.today()
    return hoy - fecha_date


def contar_doble_entrada(vehiculos):
    try:
        doble_entrada_contar = vehiculos.annotate(entrada=F("presion_de_entrada")/F("presion_de_salida")).filter(entrada__lt=0.9).values("numero_economico").annotate(count=Count("numero_economico")).values("count")
        return doble_entrada_contar
    except:
        return None

def contar_entrada_correcta(vehiculos):
    try:
        entrada_correcta_contar = vehiculos.annotate(entrada=F("presion_de_entrada")/F("presion_de_salida")).filter(entrada__gte=0.9).count()
        return entrada_correcta_contar
    except:
        return None


def contar_mala_entrada(vehiculos):
    try:
        mala_entrada_contar = vehiculos.annotate(entrada=F("presion_de_entrada")/F("presion_de_salida")).filter(entrada__lt=0.9).count()
        return mala_entrada_contar
    except:
        return None



def convertir_fecha(fecha):
    try:
        partes_fecha = fecha.split("-")
        return f"{partes_fecha[2]}/{partes_fecha[1]}/{partes_fecha[0][2:4]}"
    except:
        return None

def convertir_rango(fecha):
    partes_fecha = fecha.split("-")
    fecha_1 = f"{partes_fecha[0][0:2]}/{partes_fecha[0][3:5]}/{partes_fecha[0][6:10]}"
    fecha_2 = f"{partes_fecha[1][1:3]}/{partes_fecha[1][4:6]}/{partes_fecha[1][7:11]}"
    lista_fechas = [fecha_1, fecha_2]
    return lista_fechas
    

def cpk_vehiculo_cantidad(cpk_vehiculos):
    max_v = max(cpk_vehiculos)
    min_v = min(cpk_vehiculos)
    resta = max_v - min_v
    div = resta / 8

    valores = []
    for i in range(9):
        valores.append(round(min_v + (div * i), 4))

    rangos = []
    rango_1 = 0
    rango_2 = 0
    rango_3 = 0
    rango_4 = 0
    rango_5 = 0
    rango_6 = 0
    rango_7 = 0
    rango_8 = 0
    for cpk in cpk_vehiculos:
        if valores[0] <= cpk < valores[1]:
            rango_1 += 1
        elif valores[1] <= cpk < valores[2]:
            rango_2 += 1
        elif valores[2] <= cpk < valores[3]:
            rango_3 += 1
        elif valores[3] <= cpk < valores[4]:
            rango_4 += 1
        elif valores[4] <= cpk < valores[5]:
            rango_5 += 1
        elif valores[5] <= cpk < valores[6]:
            rango_6 += 1
        elif valores[6] <= cpk < valores[7]:
            rango_7 += 1
        elif valores[7] <= cpk <= valores[8]:
            rango_8 += 1
    rangos.append(rango_1)
    rangos.append(rango_2)
    rangos.append(rango_3)
    rangos.append(rango_4)
    rangos.append(rango_5)
    rangos.append(rango_6)
    rangos.append(rango_7)
    rangos.append(rango_8)

    return valores, rangos


def desdualizacion(llantas, periodo):
    relacion = {}
    vehiculos = []
    for llanta in llantas:
        relacion[llanta] = llanta.vehiculo
        if not(llanta.vehiculo in vehiculos):
            vehiculos.append(llanta.vehiculo)
    
    for vehiculo in vehiculos:
        posicion = {}
        dual = {}
        for key, val in relacion.items():
            if vehiculo == val:
                if key.posicion == ("1RI" or "1RO"):
                    if "0" in posicion:
                        dual[key] = posicion["0"]
                    else:
                        posicion["0"] = key
                elif key.posicion == ("1LI" or "1LO"):
                    if "1" in posicion:
                        dual[key] = posicion["1"]
                    else:
                        posicion["1"] = key
                elif key.posicion == ("2RI" or "2RO"):
                    if "2" in posicion:
                        dual[key] = posicion["2"]
                    else:
                        posicion["2"] = key
                elif key.posicion == ("2LI" or "2LO"):
                    if "3" in posicion:
                        dual[key] = posicion["3"]
                    else:
                        posicion["3"] = key
                elif key.posicion == ("4RI" or "4RO"):
                    if "4" in posicion:
                        dual[key] = posicion["4"]
                    else:
                        posicion["4"] = key
                elif key.posicion == ("4LI" or "4LO"):
                    if "5" in posicion:
                        dual[key] = posicion["5"]
                    else:
                        posicion["5"] = key

    eje = {}
    for key, val in dual.items():
        if periodo:
            profundidad_1 = key.min_profundidad
            profundidad_2 = val.min_profundidad
        else:
            profundidad_1 = key.primera_min_profundidad
            profundidad_2 = val.primera_min_profundidad
        if (profundidad_1 - profundidad_2) < -3 or (profundidad_1 - profundidad_2) > 3:
            if key.tipo_de_eje[0] == "S":
                if "direccion" in eje:
                    eje["direccion"] += 1
                else:
                    eje["direccion"] = 1
            elif key.tipo_de_eje[0] == "D":
                if "traccion" in eje:
                    eje["traccion"] += 1
                else:
                    eje["traccion"] = 1
            elif key.tipo_de_eje[0] == "T":
                if "arrastre" in eje:
                    eje["arrastre"] += 1
                else:
                    eje["arrastre"] = 1
            elif key.tipo_de_eje[0] == "C":
                if "loco" in eje:
                    eje["loco"] += 1
                else:
                    eje["loco"] = 1
            elif key.tipo_de_eje[0] == "L":
                if "retractil" in eje:
                    eje["retractil"] += 1
                else:
                    eje["retractil"] = 1

    return eje


def desgaste_irregular(llantas, periodo):
    if periodo:
        llantas_desgaste_irregular = llantas.select_related("ultima_inspeccion", "vehiculo__compania").annotate(punto_de_retiro=Case(When(nombre_de_eje="Dirección", then=F("vehiculo__compania__punto_retiro_eje_direccion")),When(nombre_de_eje="Tracción", then=F("vehiculo__compania__punto_retiro_eje_traccion")),When(nombre_de_eje="Arrastre", then=F("vehiculo__compania__punto_retiro_eje_arrastre")),When(nombre_de_eje="Loco", then=F("vehiculo__compania__punto_retiro_eje_loco")),When(nombre_de_eje="Retractil", then=F("vehiculo__compania__punto_retiro_eje_retractil")))).values("numero_economico").filter(punto_de_retiro__lt=F("ultima_inspeccion__max_profundidad")-F("ultima_inspeccion__min_profundidad"))
    else:
        llantas_desgaste_irregular = llantas.select_related("ultima_inspeccion", "vehiculo__compania").annotate(llanta_eje=Substr(F("tipo_de_eje"),1,1)).annotate(punto_de_retiro=Case(When(llanta_eje="S", then=F("vehiculo__compania__punto_retiro_eje_direccion")),When(llanta_eje="D", then=F("vehiculo__compania__punto_retiro_eje_traccion")),When(llanta_eje="T", then=F("vehiculo__compania__punto_retiro_eje_arrastre")),When(llanta_eje="C", then=F("vehiculo__compania__punto_retiro_eje_loco")),When(llanta_eje="L", then=F("vehiculo__compania__punto_retiro_eje_retractil")))).filter(punto_de_retiro__gt=F("ultima_inspeccion__max_profundidad")-F("ultima_inspeccion__min_profundidad")).annotate(nombre_eje=Case(When(llanta_eje="S", then=Value("direccion")),When(llanta_eje="D", then=Value("traccion")),When(llanta_eje="T", then=Value("arrastre")),When(llanta_eje="C", then=Value("loco")),When(llanta_eje="L", then=Value("retractil")), output_field=CharField())).values("nombre_eje").annotate(total=Count("nombre_eje")).values("nombre_eje", "total")
    llantas_desgaste_ejes = llantas_desgaste_irregular.aggregate(direccion=Count("nombre_de_eje",filter=Q(nombre_de_eje="Dirección")), traccion=Count("nombre_de_eje",filter=Q(nombre_de_eje="Tracción")), arrastre=Count("nombre_de_eje",filter=Q(nombre_de_eje="Arrastre")), loco=Count("nombre_de_eje",filter=Q(nombre_de_eje="Loco")), retractil=Count("nombre_de_eje",filter=Q(nombre_de_eje="Retractil")))
    desgaste_ejes = {k: v for k, v in llantas_desgaste_ejes.items() if v != 0}
    return desgaste_ejes

def desgaste_irregular_producto(llantas):
    llantas_desgaste_irregular = llantas.select_related("ultima_inspeccion", "vehiculo__compania").annotate(punto_de_retiro=Case(When(nombre_de_eje="Dirección", then=F("vehiculo__compania__punto_retiro_eje_direccion")),When(nombre_de_eje="Tracción", then=F("vehiculo__compania__punto_retiro_eje_traccion")),When(nombre_de_eje="Arrastre", then=F("vehiculo__compania__punto_retiro_eje_arrastre")),When(nombre_de_eje="Loco", then=F("vehiculo__compania__punto_retiro_eje_loco")),When(nombre_de_eje="Retractil", then=F("vehiculo__compania__punto_retiro_eje_retractil")))).values("numero_economico").filter(punto_de_retiro__lt=F("ultima_inspeccion__max_profundidad")-F("ultima_inspeccion__min_profundidad")).count()
    resta = llantas.count() - llantas_desgaste_irregular
    porcentaje = round((resta / llantas.count()) * 100, 2)
    return porcentaje


def distribucion_cantidad(cpks):
    for cpk in cpks:
        min = np.quantile(cpks[cpk], 0)
        q1 = np.quantile(cpks[cpk], 0.25)
        q2 = np.quantile(cpks[cpk], 0.5)
        q3 = np.quantile(cpks[cpk], 0.75)
        max = np.quantile(cpks[cpk], 1)
        
        cpks[cpk] = [min, q1, q2, q3, max]

    return cpks


def doble_entrada(bitacoras):
    bitacora = bitacoras.annotate(presiones_de_entrada=Cast(F("presion_de_entrada"), FloatField()) / Cast(F("presion_de_salida"), FloatField())).filter(presiones_de_entrada__lte=0.9).values("numero_economico").annotate(max=Max("fecha_de_inflado")).annotate(count=Count("numero_economico")).filter(count__gt=1).annotate(mes=(ExtractYear(Now()) - ExtractYear("max")) * 12 + (ExtractMonth(Now()) - ExtractMonth("max")) + 1)
    meses = bitacora.values("mes").aggregate(mes1=Count("mes",filter=Q(mes=1),distinct=True), mes2=Count("mes",filter=Q(mes=2),distinct=True), mes3=Count("mes",filter=Q(mes=3),distinct=True), mes4=Count("mes",filter=Q(mes=4),distinct=True))
    bitacora = bitacora.values("numero_economico")
    return bitacora, meses

def doble_mala_entrada(bitacoras, vehiculos):
    bitacora = bitacoras.annotate(presiones_de_entrada=Cast(F("presion_de_entrada"), FloatField()) / Cast(F("presion_de_salida"), FloatField())).filter(presiones_de_entrada__lte=0.9).values("numero_economico").annotate(max=Max("fecha_de_inflado")).annotate(count=Count("numero_economico")).filter(count__gt=1).values("numero_economico")
    try:
        vehiculos = vehiculos.filter(id__in=bitacora)
        return vehiculos
    except:
        if bitacora.exists():
            return vehiculos
        return None

def duales(llantas):
    dual = llantas.values("numero_economico", "vehiculo", "posicion").annotate(eje_dual=Substr(F("posicion"),1,2)).values("vehiculo", "eje_dual").order_by().annotate(Count("numero_economico"), llanta_1=Max("numero_economico"), llanta_2=Min("numero_economico")).filter(numero_economico__count=2)
    dual_1 = dual.values_list("llanta_1", flat=True)
    dual_2 = dual.values_list("llanta_2", flat=True)
    llantas_dual_1 = llantas.filter(numero_economico__in=dual_1)
    llantas_dual_2 = llantas.filter(numero_economico__in=dual_2)
    zip_llantas = zip(dual_1, dual_2)
    llantas_dictionary = dict(zip_llantas)
    return llantas_dual_1, llantas_dual_2, llantas_dictionary

def embudo_vidas(llantas):
    llantas_abajo_del_punto = llantas.select_related("vehiculo__compania", "vehiculo__ubicacion").annotate(llanta_eje=Substr(F("tipo_de_eje"),1,1)).annotate(punto_de_retiro=Case(When(llanta_eje="S", then=F("vehiculo__compania__punto_retiro_eje_direccion")),When(llanta_eje="D", then=F("vehiculo__compania__punto_retiro_eje_traccion")),When(llanta_eje="T", then=F("vehiculo__compania__punto_retiro_eje_arrastre")),When(llanta_eje="C", then=F("vehiculo__compania__punto_retiro_eje_loco")),When(llanta_eje="L", then=F("vehiculo__compania__punto_retiro_eje_retractil")))).filter(ultima_inspeccion__min_profundidad__lte=F("punto_de_retiro"))
    llantas_vida = llantas_abajo_del_punto.aggregate(nueva=Count("vida",filter=Q(vida="Nueva")), r1=Count("vida",filter=Q(vida="1R")), r2=Count("vida",filter=Q(vida="2R")), r3=Count("vida",filter=Q(vida="3R")), r4=Count("vida",filter=Q(vida="4R")), r5=Count("vida",filter=Q(vida="5R")), total=Count("vida"))
    renovado = llantas_abajo_del_punto.annotate(renovado1=Round(Cast(F("vehiculo__ubicacion__rendimiento_de_nueva"),FloatField()) * Cast(llantas_vida["nueva"],FloatField()) / 100), renovado2=Round(Cast(F("vehiculo__ubicacion__rendimiento_de_primera"),FloatField()) * Cast(llantas_vida["r1"],FloatField()) / 100), renovado3=Round(Cast(F("vehiculo__ubicacion__rendimiento_de_segunda"),FloatField()) * Cast(llantas_vida["r2"],FloatField()) / 100), renovado4=Round(Cast(F("vehiculo__ubicacion__rendimiento_de_tercera"),FloatField()) * Cast(llantas_vida["r3"],FloatField()) / 100), renovado5=Round(Cast(F("vehiculo__ubicacion__rendimiento_de_cuarta"),FloatField()) * Cast(llantas_vida["r4"],FloatField()) / 100)).annotate(renovadonuevo=llantas_vida["total"]-F("renovado1")-F("renovado2")-F("renovado3")-F("renovado4")-F("renovado5")).values("renovadonuevo", "renovado1", "renovado2", "renovado3", "renovado4", "renovado5").distinct()
    if not(llantas_abajo_del_punto):
        llantas_abajo_del_punto = 0
    if renovado:
        return llantas_abajo_del_punto, renovado[0]
    else:
        return llantas_abajo_del_punto, {'renovado1': 0, 'renovado2': 0, 'renovado3': 0, 'renovado4': 0, 'renovado5': 0, 'renovadonuevo': 0}

def embudo_vidas_con_regresion(inspecciones):
    vehiculos_sospechosos = []
    """triplicadas = inspecciones.select_related("llanta").values("llanta").annotate(count=Count("llanta")).filter(count__gt=2)
    regresion = inspecciones.select_related("llanta__vehiculo").annotate(poli=Case(When(llanta__in=triplicadas.values("llanta"), then=1), default=0, output_field=IntegerField())).filter(poli=1)
    llantas = regresion.values("llanta").distinct()
    for llanta in llantas:
        x = []
        y = []
        primera_fecha = regresion.filter(llanta=llanta["llanta"]).aggregate(primera_fecha=Min("fecha_hora"))
        fecha_llanta = regresion.filter(llanta=llanta["llanta"]).values("fecha_hora")
        for r in fecha_llanta:
            resta = abs(r["fecha_hora"] - primera_fecha["primera_fecha"]).days
            x.append(resta)
        profundidades = regresion.filter(llanta=llanta["llanta"]).values("min_profundidad")
        for p in profundidades:
            y.append(p["min_profundidad"])
        
        x = np.array(x)
        y = np.array(y)

        if len(x) > 2:
            dia = x[-1]

            f = np.polyfit(x, y, 3)
            p = np.poly1d(f)
            termino = []
            for numero in p:
                numero = round(numero, 4)
                termino.append(numero)
            regresion_resultado = (termino[0]*(dia**2))+(termino[1]*dia)+termino[2]
            resta = y[0]-regresion_resultado
            diario = resta/dia
            dias_30 = resta - (diario * 30)
            dias_60 = resta - diario * 60
            dias_90 = resta - diario * 90
            
            #vehiculos_sospechosos = regresion.filter(min_profundidad__gt=desgaste_normal).values("llanta__vehiculo").distinct()"""
    duplicadas = inspecciones.select_related("llanta").values("llanta").annotate(count=Count("llanta")).filter(count__lte=2, count__gt=0)
    sin_regresion = inspecciones.select_related("llanta__vehiculo__compania").annotate(poli=Case(When(llanta__in=duplicadas.values("llanta"), then=1), default=0, output_field=IntegerField())).filter(poli=1)
    llantas = sin_regresion.values("llanta").distinct()
    vehiculos_lista = []
    if vehiculos_sospechosos:
        vehiculos_sospechosos_iteracion = True
    else:
        vehiculos_sospechosos_iteracion = False
    for llanta in llantas:
        fechas = sin_regresion.filter(llanta=llanta["llanta"]).aggregate(primera_fecha=Min("fecha_hora"),ultima_fecha=Max("fecha_hora"))
        dias = abs(fechas["ultima_fecha"] - fechas["primera_fecha"]).days
        if dias == 0:
            dias = 1
        print(dias)
        resta = sin_regresion.filter(llanta=llanta["llanta"]).order_by("fecha_hora").first().min_profundidad - sin_regresion.filter(llanta=llanta["llanta"]).order_by("fecha_hora").last().min_profundidad
        print(resta)
        diario = resta/dias
        print(diario)
        dias_30 = sin_regresion.filter(llanta=llanta["llanta"]).order_by("fecha_hora").last().min_profundidad - (diario * 30)
        dias_60 = sin_regresion.filter(llanta=llanta["llanta"]).order_by("fecha_hora").last().min_profundidad - (diario * 60)
        dias_90 = sin_regresion.filter(llanta=llanta["llanta"]).order_by("fecha_hora").last().min_profundidad - (diario * 90)

        print("dias_30: ", dias_30)
        print("dias_60: ", dias_60)
        print("dias_90: ", dias_90)

        vehiculo = sin_regresion.filter(llanta=llanta["llanta"]).filter(min_profundidad__lt = resta - (F("llanta__vehiculo__compania__mm_parametro_sospechoso") * dias)).values("llanta__vehiculo").distinct()
        if vehiculo:
            if not(vehiculo[0]["llanta__vehiculo"] in vehiculos_lista):
                if vehiculos_sospechosos_iteracion:
                    vehiculos_sospechosos = vehiculos_sospechosos.union(vehiculo)
                else:
                    vehiculos_sospechosos = vehiculo
                    vehiculos_sospechosos_iteracion = True
                vehiculos_lista.append(vehiculo[0]["llanta__vehiculo"])
    return vehiculos_sospechosos

def entrada_correcta(vehiculos):
    try:
        entradas = {}
        for vehiculo in vehiculos:
            presion_encontrada = vehiculo.presion_de_entrada
            presion_establecida = vehiculo.presion_de_salida
            entrada_correcta = presion_encontrada/presion_establecida

            if entrada_correcta >= 0.9:
                entradas[vehiculo.id] = True
            else:
                entradas[vehiculo.id] = False
        return entradas
    except:
        try:
            presion_encontrada = vehiculos.presion_de_entrada
            presion_establecida = vehiculos.presion_de_salida
            entrada_correcta = presion_encontrada/presion_establecida

            if entrada_correcta >= 0.9:
                entradas = "good"
            else:
                entradas = "bad"
            return entradas
        except:
            return None

def estatus_profundidad(llantas):
    estatus = llantas.select_related("ultima_inspeccion", "vehiculo__compania").annotate(llanta_eje=Substr(F("tipo_de_eje"),1,1)).annotate(punto_de_retiro=Case(When(llanta_eje="S", then=F("vehiculo__compania__punto_retiro_eje_direccion")),When(llanta_eje="D", then=F("vehiculo__compania__punto_retiro_eje_traccion")),When(llanta_eje="T", then=F("vehiculo__compania__punto_retiro_eje_arrastre")),When(llanta_eje="C", then=F("vehiculo__compania__punto_retiro_eje_loco")),When(llanta_eje="L", then=F("vehiculo__compania__punto_retiro_eje_retractil"))), nombre_eje=Case(When(llanta_eje="S", then=Value("direccion")),When(llanta_eje="D", then=Value("traccion")),When(llanta_eje="T",then=Value("arrastre")),When(llanta_eje="C", then=Value("loco")),When(llanta_eje="L", then=Value("retractil")), output_field=CharField())).annotate(estatus=Case(When(ultima_inspeccion__min_profundidad__gt=F("punto_de_retiro"), then=Value("verde")),When(ultima_inspeccion__min_profundidad__gte=F("punto_de_retiro"),then=Value("amarillo")),When(ultima_inspeccion__min_profundidad__lte=F("punto_de_retiro"),then=Value("rojo")), output_field=CharField())).values("nombre_eje").annotate(num_verde=Count("estatus",filter=Q(estatus="verde")),num_amarillo=Count("estatus",filter=Q(estatus="amarillo")),num_rojo=Count("estatus",filter=Q(estatus="rojo")))
    return estatus

def inflado_promedio(vehiculo):
    tiempo_promedio = 0
    promedio_contar = 0
    for tiempo in vehiculo:
        try:
            tiempo_promedio += tiempo.tiempo_de_inflado
            promedio_contar +=1
        except:
            pass
    try:
        return round(tiempo_promedio/promedio_contar, 2)
    except:
        return None

def km_proyectado(inspecciones, mediana):
    duplicadas = inspecciones.select_related("llanta").values("llanta").annotate(count=Count("llanta")).filter(count__gt=1)
    regresion = inspecciones.select_related("llanta__vehiculo").annotate(poli=Case(When(llanta__in=duplicadas.values("llanta"), then=1), default=0, output_field=IntegerField())).filter(poli=1)
    llantas = regresion.values("llanta").distinct()
    llantas_limpias = []
    kms_proyectados = []
    kms_x_mm = []
    cpks = []
    for llanta in llantas:
        llanta_completa = Llanta.objects.get(id=llanta["llanta"])
        if llanta_completa.km_montado:
            x = []
            y = []
            x.append(llanta_completa.km_montado)
            y.append(llanta_completa.producto.profundidad_inicial)
            km_llanta = regresion.filter(llanta=llanta["llanta"]).values("km")
            for r in km_llanta:
                suma = abs(r["km"] + llanta_completa.km_montado)
                y.append(suma)
            profundidades = regresion.filter(llanta=llanta["llanta"]).values("min_profundidad")
            for p in profundidades:
                x.append(p["min_profundidad"])
            
            x = np.array(x)
            y = np.array(y)

            f = np.polyfit(x, y, 2)
            p = np.poly1d(f)
            termino = []
            for numero in p:
                numero = round(numero, 4)
                termino.append(numero)
            km_actual = y[-1]
            km_proyectado = (termino[0]*(3**2))+(termino[1]*3)+termino[2]
            km_x_mm = km_proyectado / (llanta_completa.producto.profundidad_inicial - 3)
            precio = llanta_completa.producto.precio
            cpk = (precio / km_proyectado)

            if km_actual >= 20000:

                llantas_limpias.append(llanta_completa)
                kms_proyectados.append(km_proyectado)
                kms_x_mm.append(km_x_mm)
                cpks.append(cpk)

        else:
            valores_llanta = regresion.filter(llanta=llanta["llanta"]).aggregate(max_mm=Max("min_profundidad"), km_recorrido=Max("km")-Min("km"), min_mm=Min("min_profundidad"))
            mm_desgastados = valores_llanta["max_mm"] - valores_llanta["min_mm"]
            if mm_desgastados == 0:
                mm_desgastados = 1
            km_recorrido = valores_llanta["km_recorrido"]
            km_x_mm = km_recorrido / mm_desgastados
            profundidad_inicial = llanta_completa.producto.profundidad_inicial
            km_teorico_actual = int((profundidad_inicial - valores_llanta["min_mm"]) * km_x_mm)
            km_teorico_proyectado = int((profundidad_inicial - 3) * km_x_mm)

            precio = llanta_completa.producto.precio
            cpk = (precio / km_teorico_proyectado)

            """print("llanta", llanta)
            print("mm_desgastados", mm_desgastados)
            print("km_recorrido", km_recorrido)
            print("km_x_mm", km_x_mm)
            print("profundidad_inicial", profundidad_inicial)
            print("km_teorico_actual", km_teorico_actual)
            print("km_proyectado", km_teorico_proyectado)
            print("cpk", cpk)"""

            if km_teorico_actual >= 20000:

                llantas_limpias.append(llanta_completa)
                kms_proyectados.append(km_teorico_proyectado)
                kms_x_mm.append(km_x_mm)
                cpks.append(cpk)
        
    """print("km proyectados: ", kms_proyectados)
    print("cpks: ", cpks)"""
    
    if mediana:
        mediana_km_proyectado = km_proyectado_mediana(kms_proyectados)
    else:
        try:
            mediana_km_proyectado = int(statistics.median(kms_proyectados))
        except:
            mediana_km_proyectado = 0
    try:
        mediana_kms_x_mm = int(statistics.median(kms_x_mm))
        print(mediana_kms_x_mm)
    except:
        mediana_kms_x_mm = 0
    try:
        mediana_cpks = round(statistics.median(cpks), 3)
    except:
        mediana_cpks = 0
    return mediana_km_proyectado, mediana_kms_x_mm, mediana_cpks, cpks, llantas_limpias

def km_proyectado_mediana(kms_proyectados):
    q1 =  np.quantile(kms_proyectados, 0.25)
    primer_mediana = round(statistics.median(kms_proyectados), 2)
    q3 =  np.quantile(kms_proyectados, 0.75)
    sigma = round(statistics.pstdev(kms_proyectados), 2)
    
    nuevos_kms_proyectados = []
    for km in kms_proyectados:
        if q1 <= km <= q3:
            nuevos_kms_proyectados.append(km)

    segunda_mediana = round(statistics.median(nuevos_kms_proyectados), 2)
    desvest = round(statistics.pstdev(nuevos_kms_proyectados), 2)
    limite_inferior = round(segunda_mediana - (1.5 * desvest), 2)
    limite_superior = round(segunda_mediana + (1.5 * desvest), 2)

    kms_proyectados_final = []
    for km in nuevos_kms_proyectados:
        if limite_inferior <= km <= limite_superior:
            kms_proyectados_final.append(km)

    mediana_final = int(statistics.median(kms_proyectados_final))

    """print("kms_proyectados", kms_proyectados)
    print()
    print("q1", q1)
    print("primer_mediana", primer_mediana)
    print("q3", q3)
    print("sigma", sigma)
    print()
    print("nuevos_kms_proyectados", nuevos_kms_proyectados)
    print()
    print("limite_inferior", limite_inferior)
    print("segunda_mediana", segunda_mediana)
    print("limite_superior", limite_superior)
    print("desvest", desvest)
    print()
    print("kms_proyectados_final", kms_proyectados_final)
    print("mediana_final", mediana_final)"""

    return mediana_final


def mala_entrada(vehiculos):
    vehiculos_fallidos = {}
    vehiculos_fallidos = vehiculos.annotate(entrada=Cast(F("presion_de_entrada"),FloatField())/Cast(F("presion_de_salida"),FloatField())).filter(entrada__lt=0.9)
    return vehiculos_fallidos


def mes_anterior(fecha):
    primer_dia = fecha.replace(day=1)
    return primer_dia - timedelta(days=1)

def nunca_vistos(vehiculos):
    nunca_visto = vehiculos.filter(ultima_inspeccion__fecha_hora__isnull=True).count()
    return nunca_visto


def porcentaje(divisor, dividendo):
    try:
        return int((divisor/dividendo)*100)
    except:
        return "Nada"

def presion_llantas(llantas, periodo):
    if periodo:
        llantas_presion_actual = llantas.select_related("vehiculo__compania").filter(vehiculo__compania__objetivo__lt=F("presion_de_salida") - F("presion_de_entrada")).annotate(llanta_eje=Substr(F("tipo_de_eje"),1,1)).annotate(nombre_eje=Case(When(llanta_eje="S", then=Value("direccion")),When(llanta_eje="D", then=Value("traccion")),When(llanta_eje="T", then=Value("arrastre")),When(llanta_eje="C", then=Value("loco")),When(llanta_eje="L", then=Value("retractil")),output_field=CharField())).values("nombre_eje").annotate(total=Count("nombre_eje")).values("nombre_eje", "total")
        return llantas_presion_actual
    else:
        llantas_presion_encontrada = llantas.select_related("vehiculo__compania").filter(vehiculo__compania__objetivo__lt=F("presion_de_salida") - F("presion_de_entrada")).annotate(llanta_eje=Substr(F("tipo_de_eje"),1,1)).annotate(nombre_eje=Case(When(llanta_eje="S", then=Value("direccion")),When(llanta_eje="D", then=Value("traccion")),When(llanta_eje="T", then=Value("arrastre")),When(llanta_eje="C", then=Value("loco")),When(llanta_eje="L", then=Value("retractil")),output_field=CharField())).values("nombre_eje").annotate(total=Count("nombre_eje")).values("nombre_eje", "total")
        return llantas_presion_encontrada

def presupuesto(vidas, ubicacion):
    try:
        presupuesto_r1 = 0
        presupuesto_r2 = 0
        presupuesto_r3 = 0
        presupuesto_r4 = 0
        presupuesto_r5 = 0
        presupuesto_nuevo = 0
        if "renovado1" in vidas:
            presupuesto_r1 = vidas["renovado1"]*ubicacion.precio_renovada
        if "renovado2" in vidas:
            presupuesto_r2 = vidas["renovado2"]*ubicacion.precio_renovada
        if "renovado3" in vidas:
            presupuesto_r3 = vidas["renovado3"]*ubicacion.precio_renovada
        if "renovado4" in vidas:
            presupuesto_r4 = vidas["renovado4"]*ubicacion.precio_renovada
        if "renovado5" in vidas:
            presupuesto_r5 = vidas["renovado5"]*ubicacion.precio_renovada
        if "renovadonuevo" in vidas:
            presupuesto_nuevo = vidas["renovadonuevo"]*ubicacion.precio_nueva
        presupuesto_total = presupuesto_r1 + presupuesto_r2 + presupuesto_r3 + presupuesto_r4 + presupuesto_r5 + presupuesto_nuevo
        return presupuesto_total
    except:
        return 0

def pronostico_de_consumo(embudo):
    try:
        ejes = embudo.aggregate(direccion=Count("nombre_de_eje",filter=Q(nombre_de_eje="Dirección")), traccion=Count("nombre_de_eje",filter=Q(nombre_de_eje="Tracción")), arrastre=Count("nombre_de_eje",filter=Q(nombre_de_eje="Arrastre")), loco=Count("nombre_de_eje",filter=Q(nombre_de_eje="Loco")), retractil=Count("nombre_de_eje",filter=Q(nombre_de_eje="Retractil")))
        return ejes
    except:
        return {'direccion': 0, 'traccion': 0, 'arrastre': 0, 'loco': 0, 'retractil': 0}

def radar_min(vehiculo_fecha, compania):
    try:
        clases = clases_mas_frecuentes(vehiculo_fecha, compania)
        clases_min = min(clases.keys(), key=lambda k: clases[k])
        return clases[clases_min]
    except:
        return None

def radar_max(vehiculo_fecha, compania):
    try:
        clases = clases_mas_frecuentes(vehiculo_fecha, compania)
        clases_max = max(clases.keys(), key=lambda k: clases[k])
        return clases[clases_max]
    except:
        return None

def radar_min_resta(min, max):
    try:
        resta = max - min
        resta = resta / 2.3
        if resta == 0:
            resta = 1
        return min - resta
    except:
        return None

def reemplazo_actual(llantas):
    reemplazo_actual_llantas = llantas.select_related("ultima_inspeccion").annotate(punto_de_retiro=Case(When(nombre_de_eje="Dirección", then=F("vehiculo__compania__punto_retiro_eje_direccion")),When(nombre_de_eje="Tracción", then=F("vehiculo__compania__punto_retiro_eje_traccion")),When(nombre_de_eje="Arrastre", then=F("vehiculo__compania__punto_retiro_eje_arrastre")),When(nombre_de_eje="Loco", then=F("vehiculo__compania__punto_retiro_eje_loco")),When(nombre_de_eje="Retractil", then=F("vehiculo__compania__punto_retiro_eje_retractil")))).filter(ultima_inspeccion__min_profundidad__lte=F("punto_de_retiro"))
    reemplazo_actual_ejes = reemplazo_actual_llantas.aggregate(direccion=Count("nombre_de_eje",filter=Q(nombre_de_eje="Dirección")),traccion=Count("nombre_de_eje",filter=Q(nombre_de_eje="Tracción")),arrastre=Count("nombre_de_eje",filter=Q(nombre_de_eje="Arrastre")),loco=Count("nombre_de_eje",filter=Q(nombre_de_eje="Loco")),retractil=Count("nombre_de_eje",filter=Q(nombre_de_eje="Retractil")),total=Count("nombre_de_eje"))
    return reemplazo_actual_llantas, reemplazo_actual_ejes

def reemplazo_dual(llantas, reemplazo_actual):
    try:
        llantas_duales = duales(llantas)
        reemplazo_dual_1 = llantas_duales[0].filter(id__in=reemplazo_actual)
        reemplazo_dual_2 = llantas_duales[1].filter(id__in=reemplazo_actual)
        dual_dictionary = llantas_duales[2]
        reemplazo_dual_1_list = reemplazo_dual_1.values_list("numero_economico", flat=True)
        reemplazo_dual_2_list = reemplazo_dual_2.values_list("numero_economico", flat=True)
        array_of_qs = []
        for k, v in dual_dictionary.items():
            array_of_qs.append(reemplazo_dual_1.filter(numero_economico=k).annotate(pareja=ExpressionWrapper(Value(v),output_field=CharField())).exclude(pareja__in=reemplazo_dual_2_list).values("id"))
        for k, v in dual_dictionary.items():
            array_of_qs.append(reemplazo_dual_2.filter(numero_economico=v).annotate(pareja=ExpressionWrapper(Value(k),output_field=CharField())).exclude(pareja__in=reemplazo_dual_1_list).values("id"))
        reemplazo_dual = array_of_qs[0].union(*array_of_qs[1:])
        reemplazo_dual_llantas = llantas.filter(id__in=reemplazo_dual)
        reemplazo_dual_ejes = reemplazo_dual_llantas.aggregate(direccion=Count("nombre_de_eje",filter=Q(nombre_de_eje="Dirección")),traccion=Count("nombre_de_eje",filter=Q(nombre_de_eje="Tracción")),arrastre=Count("nombre_de_eje",filter=Q(nombre_de_eje="Arrastre")),loco=Count("nombre_de_eje",filter=Q(nombre_de_eje="Loco")),retractil=Count("nombre_de_eje",filter=Q(nombre_de_eje="Retractil")),total=Count("nombre_de_eje"))
        reemplazo_dual_ejes = {k: v for k, v in reemplazo_dual_ejes.items() if v != 0}
        return reemplazo_dual_ejes
    except:
        return None

def reemplazo_total(reemplazo_actual, reemplazo_dual):
    cantidad_total = {}
    try:
        for reemplazo in reemplazo_actual:
            r = reemplazo_actual[reemplazo]
            cantidad_total[reemplazo] = r
        for reemplazo in reemplazo_dual:
            if not(reemplazo in cantidad_total):
                cantidad_total[reemplazo] = reemplazo_dual[reemplazo]
            else:
                cantidad_total[reemplazo] += reemplazo_dual[reemplazo]
        
        return cantidad_total
    except:
        return None

def renovables(llantas, vehiculos_amarillos):
    cantidad_renovables = llantas.select_related("vehiculo").filter(vehiculo__in=vehiculos_amarillos).count()
    return cantidad_renovables

def salida_correcta(vehiculos):
    try:
        presion_de_salida = vehiculos.presion_de_salida
        presion_establecida = vehiculos.presion_establecida
        entrada_correcta = presion_de_salida/presion_establecida

        if entrada_correcta >= 0.9:
            entradas = "good"
        else:
            entradas = "bad"
        return entradas
    except:
        return None

def sin_informacion(llantas):
    llantas_sin_informacion = llantas.filter(producto__isnull=True).count()
    return llantas_sin_informacion

def vehiculo_amarillo(llantas):
    vehiculos_amarillos = []
    try:
        # Desgaste irregular
        llantas_desgaste = llantas.select_related("ultima_inspeccion", "vehiculo").annotate(resta=F("ultima_inspeccion__max_profundidad")-F("ultima_inspeccion__min_profundidad")).filter(resta__gt=3).values("vehiculo").distinct()

        # Arriba del punto de retiro
        llantas_retiro = llantas.select_related("ultima_inspeccion","vehiculo__compania").annotate(llanta_eje=Substr(F("tipo_de_eje"),1,1)).annotate(punto_de_retiro=Case(When(llanta_eje="S", then=F("vehiculo__compania__punto_retiro_eje_direccion")),When(llanta_eje="D", then=F("vehiculo__compania__punto_retiro_eje_traccion")),When(llanta_eje="T", then=F("vehiculo__compania__punto_retiro_eje_arrastre")),When(llanta_eje="C", then=F("vehiculo__compania__punto_retiro_eje_loco")),When(llanta_eje="L", then=F("vehiculo__compania__punto_retiro_eje_retractil")))).filter(ultima_inspeccion__min_profundidad__gt=F("punto_de_retiro") + 1).values("vehiculo").distinct()
        vehiculos_amarillos = llantas_desgaste.union(llantas_retiro)
        
        # Desdualización
        dual = llantas.select_related("ultima_inspeccion", "vehiculo").values("numero_economico", "vehiculo", "posicion").annotate(eje_dual=Substr(F("posicion"),1,2)).values("vehiculo", "eje_dual").order_by().annotate(Count("numero_economico"), diferencia=Max("ultima_inspeccion__min_profundidad") - Min("ultima_inspeccion__min_profundidad")).filter(numero_economico__count=2).filter(diferencia__gt=3).values("vehiculo").distinct()
        vehiculos_amarillos = vehiculos_amarillos.union(dual)
        return vehiculos_amarillos
    except:
        return []

def vehiculos_inspeccionados_por_aplicacion(vehiculos):
    porcentaje_aplicacion = vehiculos.select_related("ultima_inspeccion", "aplicacion").values("aplicacion__nombre").distinct().annotate(ins=Case(When(ultima_inspeccion__fecha_hora__gte=timezone.now()-timedelta(days=31),then=1),default=0,output_field=IntegerField())).annotate(inspeccionado=Count("ins",filter=Q(ins=1)),no_inspeccionado=Count("ins",filter=Q(ins=0))).values("aplicacion__nombre","inspeccionado").annotate(total=F("inspeccionado")+F("no_inspeccionado")).values("aplicacion__nombre","inspeccionado","total")
    return porcentaje_aplicacion

def vehiculos_inspeccionados_por_clase(vehiculos):
    porcentaje_clase = vehiculos.select_related("ultima_inspeccion").values("clase").distinct().annotate(ins=Case(When(ultima_inspeccion__fecha_hora__gte=timezone.now()-timedelta(days=31),then=1),default=0,output_field=IntegerField())).annotate(inspeccionado=Count("ins",filter=Q(ins=1)),no_inspeccionado=Count("ins",filter=Q(ins=0))).values("clase","inspeccionado").annotate(total=F("inspeccionado")+F("no_inspeccionado")).values("clase","inspeccionado","total")
    return porcentaje_clase

def vehiculos_por_aplicacion_sin_inspeccionar(vehiculos, hoy1, hoy2, hoy3):
    vehiculos_vencidos = vehiculos.filter(ultima_inspeccion__fecha_hora__month=hoy1) | vehiculos.filter(ultima_inspeccion__fecha_hora__month=hoy2) | vehiculos.filter(ultima_inspeccion__fecha_hora__month=hoy3)
    vehiculos_aplicacion = vehiculos_vencidos.values("aplicacion__nombre").distinct().annotate(mes1=Count("ultima_inspeccion__fecha_hora__month",filter=Q(ultima_inspeccion__fecha_hora__month=hoy1)),mes2=Count("ultima_inspeccion__fecha_hora__month",filter=Q(ultima_inspeccion__fecha_hora__month=hoy2)),mes3=Count("ultima_inspeccion__fecha_hora__month",filter=Q(ultima_inspeccion__fecha_hora__month=hoy3)))
    return vehiculos_aplicacion

def vehiculos_por_clase_sin_inspeccionar(vehiculos, hoy1, hoy2, hoy3):
    vehiculos_vencidos = vehiculos.filter(ultima_inspeccion__fecha_hora__month=hoy1) | vehiculos.filter(ultima_inspeccion__fecha_hora__month=hoy2) | vehiculos.filter(ultima_inspeccion__fecha_hora__month=hoy3)
    vehiculos_clase = vehiculos_vencidos.values("clase").distinct().annotate(mes1=Count("ultima_inspeccion__fecha_hora__month",filter=Q(ultima_inspeccion__fecha_hora__month=hoy1)),mes2=Count("ultima_inspeccion__fecha_hora__month",filter=Q(ultima_inspeccion__fecha_hora__month=hoy2)),mes3=Count("ultima_inspeccion__fecha_hora__month",filter=Q(ultima_inspeccion__fecha_hora__month=hoy3)))
    return vehiculos_clase

def vehiculo_rojo(llantas, doble_entrada):
    llantas = llantas.select_related("ultima_inspeccion","vehiculo__compania").annotate(llanta_eje=Substr(F("tipo_de_eje"),1,1)).annotate(punto_de_retiro=Case(When(llanta_eje="S", then=F("vehiculo__compania__punto_retiro_eje_direccion")),When(llanta_eje="D", then=F("vehiculo__compania__punto_retiro_eje_traccion")),When(llanta_eje="T", then=F("vehiculo__compania__punto_retiro_eje_arrastre")),When(llanta_eje="C", then=F("vehiculo__compania__punto_retiro_eje_loco")),When(llanta_eje="L", then=F("vehiculo__compania__punto_retiro_eje_retractil")))).filter(ultima_inspeccion__min_profundidad__lte=F("punto_de_retiro") - 1).values("vehiculo").distinct()
    union = doble_entrada.union(llantas)
    return union

def vehiculo_sospechoso(inspecciones):
    triplicadas = inspecciones.select_related("llanta").values("llanta").annotate(count=Count("llanta")).filter(count__gt=2)
    regresion = inspecciones.select_related("llanta__vehiculo").annotate(poli=Case(When(llanta__in=triplicadas.values("llanta"), then=1), default=0, output_field=IntegerField())).filter(poli=1)
    llantas = regresion.values("llanta").distinct()
    vehiculos_sospechosos = []
    for llanta in llantas:
        x = []
        y = []
        primera_fecha = regresion.filter(llanta=llanta["llanta"]).aggregate(primera_fecha=Min("fecha_hora"))
        fecha_llanta = regresion.filter(llanta=llanta["llanta"]).values("fecha_hora")
        for r in fecha_llanta:
            resta = abs(r["fecha_hora"] - primera_fecha["primera_fecha"]).days
            x.append(resta)
        profundidades = regresion.filter(llanta=llanta["llanta"]).values("min_profundidad")
        for p in profundidades:
            y.append(p["min_profundidad"])
        
        x = np.array(x)
        y = np.array(y)

        if len(x) > 2:
            dia = x[-1]

            f = np.polyfit(x, y, 2)
            p = np.poly1d(f)
            termino = []
            for numero in p:
                numero = round(numero, 4)
                termino.append(numero)
            regresion_resultado = (termino[0]*(dia**2))+(termino[1]*dia)+termino[2]
            resta = y[0]-regresion_resultado
            diario = resta/dia
            diferencia_dias = dia-x[-2]
            prediccion = diario*diferencia_dias
            desgaste_normal = prediccion*2.5
            vehiculos_sospechosos = regresion.filter(min_profundidad__gt=desgaste_normal).values("llanta__vehiculo").distinct()
    # En un futuro poner el parámetro sospechoso para cuando es 1 inspección
    duplicadas = inspecciones.select_related("llanta").values("llanta").annotate(count=Count("llanta")).filter(count=2)
    sin_regresion = inspecciones.select_related("llanta__vehiculo__compania").annotate(poli=Case(When(llanta__in=duplicadas.values("llanta"), then=1), default=0, output_field=IntegerField())).filter(poli=1)
    llantas = sin_regresion.values("llanta").distinct()
    vehiculos_lista = []
    if vehiculos_sospechosos:
        vehiculos_sospechosos_iteracion = True
    else:
        vehiculos_sospechosos_iteracion = False
    for llanta in llantas:
        fechas = sin_regresion.filter(llanta=llanta["llanta"]).aggregate(primera_fecha=Min("fecha_hora"),ultima_fecha=Max("fecha_hora"))
        dias = abs(fechas["ultima_fecha"] - fechas["primera_fecha"]).days
        #print("llanta: ", Llanta.objects.get(id=llanta["llanta"]))
        #print("dias: ", dias)
        primera_profundidad = sin_regresion.filter(llanta=llanta["llanta"]).order_by("fecha_hora").first().min_profundidad
        ultima_profundidad = sin_regresion.filter(llanta=llanta["llanta"]).order_by("fecha_hora").last().min_profundidad
        #print("primera_profundidad: ", primera_profundidad)
        #print("ultima_profundidad: ", ultima_profundidad)
        vehiculo = sin_regresion.filter(llanta=llanta["llanta"]).filter(min_profundidad__lt = primera_profundidad - (F("llanta__vehiculo__compania__mm_parametro_sospechoso") * dias)).values("llanta__vehiculo").distinct()
        if vehiculo:
            if not(vehiculo[0]["llanta__vehiculo"] in vehiculos_lista):
                if vehiculos_sospechosos_iteracion:
                    vehiculos_sospechosos = vehiculos_sospechosos.union(vehiculo)
                else:
                    vehiculos_sospechosos = vehiculo
                    vehiculos_sospechosos_iteracion = True
                vehiculos_lista.append(vehiculo[0]["llanta__vehiculo"])
    #print(vehiculos_sospechosos)
    return vehiculos_sospechosos
